"""
SAR-INVENTORY - Módulo de Reportes
==================================
Blueprint para generación de reportes con exportación a PDF y Excel.
"""

import io
from datetime import datetime, timezone
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from models import db, Equipo, Movimiento, Usuario

reportes_bp = Blueprint('reportes', __name__)


# =============================================================================
# FUNCIONES AUXILIARES INTERNAS
# =============================================================================

def _obtener_datos_reporte(tipo, filtro=''):
    """
    Obtiene los datos para un reporte según el tipo y filtro especificados.

    Args:
        tipo: Tipo de reporte ('general', 'estado', 'categoria', 'criticos', 'movimientos').
        filtro: Filtro adicional específico al tipo de reporte.

    Returns:
        Tupla (titulo, datos, es_movimientos) donde datos es una lista de objetos.
    """
    if tipo == 'general':
        titulo = 'Reporte General de Inventario'
        datos = Equipo.query.order_by(Equipo.codigo).all()
        return titulo, datos, False

    elif tipo == 'estado':
        filtro_estado = filtro if filtro else 'Bueno'
        titulo = f'Reporte de Equipos - Estado: {filtro_estado}'
        datos = Equipo.query.filter_by(estado=filtro_estado).order_by(Equipo.codigo).all()
        return titulo, datos, False

    elif tipo == 'categoria':
        filtro_cat = filtro if filtro else 'Primeros Auxilios'
        titulo = f'Reporte de Equipos - Categoría: {filtro_cat}'
        datos = Equipo.query.filter_by(categoria=filtro_cat).order_by(Equipo.codigo).all()
        return titulo, datos, False

    elif tipo == 'criticos':
        titulo = 'Reporte de Equipos Críticos (Malo o Cantidad < 3)'
        datos = Equipo.query.filter(
            db.or_(
                Equipo.estado == 'Malo',
                Equipo.cantidad < 3
            )
        ).order_by(Equipo.codigo).all()
        return titulo, datos, False

    elif tipo == 'movimientos':
        titulo = 'Reporte de Movimientos'
        query = Movimiento.query.order_by(Movimiento.fecha.desc())
        if filtro and filtro.lower() not in ('', 'todos'):
            titulo = f'Reporte de Movimientos - Acción: {filtro}'
            query = query.filter_by(accion=filtro)
        datos = query.all()
        return titulo, datos, True

    else:
        return 'Reporte', [], False


def _generar_resumen(equipos_datos):
    """Genera un resumen estadístico a partir de una lista de equipos."""
    total = len(equipos_datos)
    buenos = sum(1 for e in equipos_datos if e.estado == 'Bueno')
    regulares = sum(1 for e in equipos_datos if e.estado == 'Regular')
    malos = sum(1 for e in equipos_datos if e.estado == 'Malo')

    categorias = {}
    for e in equipos_datos:
        categorias[e.categoria] = categorias.get(e.categoria, 0) + 1

    return {
        'total': total,
        'buenos': buenos,
        'regulares': regulares,
        'malos': malos,
        'por_categoria': [{'categoria': k, 'cantidad': v} for k, v in sorted(categorias.items())]
    }


# =============================================================================
# RUTAS
# =============================================================================

@reportes_bp.route('/reportes')
@login_required
def reportes():
    """Renderiza la página de reportes."""
    return render_template('reportes.html')


@reportes_bp.route('/api/reportes/generar', methods=['POST'])
@login_required
def generar_reporte():
    """
    Genera datos de reporte según los parámetros proporcionados.
    Retorna datos en formato JSON para visualización en el frontend.
    """
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': 'No se recibieron datos.'}), 400

    tipo = data.get('tipo', 'general').strip()
    filtro = data.get('filtro', '').strip()

    titulo, datos, es_movimientos = _obtener_datos_reporte(tipo, filtro)

    # Construir filtros aplicados para mostrar en el reporte
    filtros_aplicados = f'Tipo: {tipo}'
    if filtro:
        filtros_aplicados += f', Filtro: {filtro}'

    if es_movimientos:
        datos_serializados = [m.to_dict() for m in datos]
        resumen = {
            'total': len(datos),
            'buenos': 0,
            'regulares': 0,
            'malos': 0,
            'por_categoria': []
        }
    else:
        datos_serializados = [e.to_dict() for e in datos]
        resumen = _generar_resumen(datos)

    return jsonify({
        'success': True,
        'titulo': titulo,
        'fecha_generacion': datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M'),
        'usuario': current_user.nombre_completo,
        'filtros_aplicados': filtros_aplicados,
        'total_registros': len(datos_serializados),
        'datos': datos_serializados,
        'resumen': resumen
    })


@reportes_bp.route('/api/reportes/exportar/pdf', methods=['GET'])
@login_required
def exportar_pdf():
    """
    Genera y descarga un reporte en formato PDF usando ReportLab.
    Query params: tipo, filtro.
    """
    tipo = request.args.get('tipo', 'general').strip()
    filtro = request.args.get('filtro', '').strip()

    titulo, datos, es_movimientos = _obtener_datos_reporte(tipo, filtro)

    # Crear buffer de memoria para el PDF
    buffer = io.BytesIO()

    # Configurar el documento PDF en landscape para más espacio
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=40,
        bottomMargin=40
    )

    # Estilos
    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        'TituloReporte',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=6,
        textColor=colors.HexColor('#1a3a5c')
    )
    estilo_subtitulo = ParagraphStyle(
        'SubtituloReporte',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=4,
        textColor=colors.HexColor('#555555')
    )
    estilo_info = ParagraphStyle(
        'InfoReporte',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_LEFT,
        spaceAfter=12,
        textColor=colors.HexColor('#333333')
    )

    elementos = []

    # Encabezado
    elementos.append(Paragraph('Grupo SAR - SAR-INVENTORY', estilo_titulo))
    elementos.append(Paragraph(titulo, estilo_subtitulo))
    elementos.append(Spacer(1, 6))
    elementos.append(Paragraph(
        f'Fecha de generación: {datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")} | '
        f'Generado por: {current_user.nombre_completo} | '
        f'Total de registros: {len(datos)}',
        estilo_info
    ))
    elementos.append(Spacer(1, 10))

    if es_movimientos:
        # Tabla de movimientos
        encabezados = ['#', 'Fecha', 'Usuario', 'Acción', 'Código', 'Equipo', 'Detalle']
        tabla_datos = [encabezados]

        for i, m in enumerate(datos, 1):
            tabla_datos.append([
                str(i),
                m.fecha.strftime('%d/%m/%Y %H:%M') if m.fecha else '',
                m.usuario.nombre_completo if m.usuario else 'Sistema',
                m.accion or '',
                m.equipo_codigo or '',
                m.equipo_nombre or '',
                (m.detalle[:80] + '...') if m.detalle and len(m.detalle) > 80 else (m.detalle or '')
            ])

        col_widths = [30, 90, 100, 60, 60, 120, 250]
    else:
        # Tabla de equipos
        encabezados = ['#', 'Código', 'Nombre', 'Cantidad', 'Estado', 'Categoría', 'Fecha Registro']
        tabla_datos = [encabezados]

        for i, e in enumerate(datos, 1):
            tabla_datos.append([
                str(i),
                e.codigo,
                e.nombre,
                str(e.cantidad),
                e.estado,
                e.categoria,
                e.fecha_registro.strftime('%d/%m/%Y') if e.fecha_registro else ''
            ])

        col_widths = [30, 70, 150, 60, 60, 120, 90]

    if len(tabla_datos) > 1:
        tabla = Table(tabla_datos, colWidths=col_widths, repeatRows=1)

        # Estilo de la tabla
        estilo_tabla = TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),

            # Cuerpo de datos
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Número de fila centrado
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Cantidad centrada
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),

            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#1a3a5c')),

            # Rayas alternas en filas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fa')]),
        ])

        # Resaltar filas con estado 'Malo' en rojo claro (solo para equipos)
        if not es_movimientos:
            for i, row in enumerate(tabla_datos[1:], 1):
                if row[4] == 'Malo':
                    estilo_tabla.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#ffe0e0'))
                elif row[4] == 'Regular':
                    estilo_tabla.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fff8e0'))

        tabla.setStyle(estilo_tabla)
        elementos.append(tabla)
    else:
        elementos.append(Paragraph('No se encontraron registros para los filtros aplicados.', estilo_info))

    # Resumen al final (solo para equipos)
    if not es_movimientos and datos:
        elementos.append(Spacer(1, 20))
        resumen = _generar_resumen(datos)
        resumen_texto = (
            f'<b>Resumen:</b> Total: {resumen["total"]} | '
            f'Buenos: {resumen["buenos"]} | '
            f'Regulares: {resumen["regulares"]} | '
            f'Malos: {resumen["malos"]}'
        )
        elementos.append(Paragraph(resumen_texto, estilo_info))

    # Construir el PDF con numeración de páginas
    def agregar_pie_pagina(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFont('Helvetica', 7)
        canvas_obj.setFillColor(colors.HexColor('#888888'))
        page_text = f'Página {doc_obj.page}'
        canvas_obj.drawRightString(
            landscape(letter)[0] - 30, 20, page_text
        )
        canvas_obj.drawString(
            30, 20, 'SAR-INVENTORY - Sistema de Gestión de Inventario'
        )
        canvas_obj.restoreState()

    doc.build(elementos, onFirstPage=agregar_pie_pagina, onLaterPages=agregar_pie_pagina)

    buffer.seek(0)
    fecha_archivo = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f'reporte_{tipo}_{fecha_archivo}.pdf'

    return send_file(
        buffer,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/pdf'
    )


@reportes_bp.route('/api/reportes/exportar/excel', methods=['GET'])
@login_required
def exportar_excel():
    """
    Genera y descarga un reporte en formato Excel usando openpyxl.
    Query params: tipo, filtro.
    """
    tipo = request.args.get('tipo', 'general').strip()
    filtro = request.args.get('filtro', '').strip()

    titulo, datos, es_movimientos = _obtener_datos_reporte(tipo, filtro)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Datos'

    # Estilos
    fuente_titulo = Font(name='Calibri', size=14, bold=True, color='1A3A5C')
    fuente_subtitulo = Font(name='Calibri', size=10, italic=True, color='555555')
    fuente_encabezado = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    fuente_datos = Font(name='Calibri', size=10)
    relleno_encabezado = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    relleno_alterno = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
    relleno_malo = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
    relleno_regular = PatternFill(start_color='FFF8E0', end_color='FFF8E0', fill_type='solid')
    alineacion_centro = Alignment(horizontal='center', vertical='center')
    alineacion_izquierda = Alignment(horizontal='left', vertical='center', wrap_text=True)
    borde_delgado = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Encabezado del reporte
    ws.merge_cells('A1:G1')
    celda_titulo = ws['A1']
    celda_titulo.value = 'Grupo SAR - SAR-INVENTORY'
    celda_titulo.font = fuente_titulo
    celda_titulo.alignment = alineacion_centro

    ws.merge_cells('A2:G2')
    celda_subtitulo = ws['A2']
    celda_subtitulo.value = titulo
    celda_subtitulo.font = fuente_subtitulo
    celda_subtitulo.alignment = alineacion_centro

    ws.merge_cells('A3:G3')
    celda_fecha = ws['A3']
    celda_fecha.value = (
        f'Fecha: {datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")} | '
        f'Generado por: {current_user.nombre_completo} | '
        f'Total: {len(datos)}'
    )
    celda_fecha.font = Font(name='Calibri', size=9, color='888888')
    celda_fecha.alignment = alineacion_centro

    fila_inicio = 5  # Fila donde comienzan los datos

    if es_movimientos:
        encabezados = ['#', 'Fecha', 'Usuario', 'Acción', 'Código', 'Equipo', 'Detalle']
    else:
        encabezados = ['#', 'Código', 'Nombre', 'Cantidad', 'Estado', 'Categoría', 'Fecha Registro']

    # Escribir encabezados
    for col, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=fila_inicio, column=col, value=encabezado)
        celda.font = fuente_encabezado
        celda.fill = relleno_encabezado
        celda.alignment = alineacion_centro
        celda.border = borde_delgado

    # Escribir datos
    for i, item in enumerate(datos):
        fila = fila_inicio + 1 + i
        es_par = (i % 2 == 1)

        if es_movimientos:
            valores = [
                i + 1,
                item.fecha.strftime('%d/%m/%Y %H:%M') if item.fecha else '',
                item.usuario.nombre_completo if item.usuario else 'Sistema',
                item.accion or '',
                item.equipo_codigo or '',
                item.equipo_nombre or '',
                item.detalle or ''
            ]
        else:
            valores = [
                i + 1,
                item.codigo,
                item.nombre,
                item.cantidad,
                item.estado,
                item.categoria,
                item.fecha_registro.strftime('%d/%m/%Y') if item.fecha_registro else ''
            ]

        for col, valor in enumerate(valores, 1):
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.font = fuente_datos
            celda.border = borde_delgado

            if col == 1:
                celda.alignment = alineacion_centro
            elif col in (4,) and not es_movimientos:
                celda.alignment = alineacion_centro
            else:
                celda.alignment = alineacion_izquierda

            # Aplicar colores por estado (solo para equipos)
            if not es_movimientos:
                if item.estado == 'Malo':
                    celda.fill = relleno_malo
                elif item.estado == 'Regular':
                    celda.fill = relleno_regular
                elif es_par:
                    celda.fill = relleno_alterno
            elif es_par:
                celda.fill = relleno_alterno

    # Auto-ajustar anchos de columna
    anchos_minimos = {1: 5, 2: 15, 3: 25, 4: 12, 5: 12, 6: 20, 7: 30}
    for col_num in range(1, len(encabezados) + 1):
        max_length = anchos_minimos.get(col_num, 12)
        column_letter = get_column_letter(col_num)

        for row in ws.iter_rows(min_row=fila_inicio, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = min(cell_length + 2, 60)

        ws.column_dimensions[column_letter].width = max_length

    # Hoja de resumen (solo para equipos)
    if not es_movimientos and datos:
        ws_resumen = wb.create_sheet(title='Resumen')
        resumen = _generar_resumen(datos)

        ws_resumen.merge_cells('A1:B1')
        ws_resumen['A1'].value = 'Resumen del Reporte'
        ws_resumen['A1'].font = fuente_titulo
        ws_resumen['A1'].alignment = alineacion_centro

        resumen_datos = [
            ['Métrica', 'Valor'],
            ['Total de Equipos', resumen['total']],
            ['Estado Bueno', resumen['buenos']],
            ['Estado Regular', resumen['regulares']],
            ['Estado Malo', resumen['malos']],
        ]

        for i, (label, valor) in enumerate(resumen_datos):
            fila = 3 + i
            celda_label = ws_resumen.cell(row=fila, column=1, value=label)
            celda_valor = ws_resumen.cell(row=fila, column=2, value=valor)
            celda_label.border = borde_delgado
            celda_valor.border = borde_delgado

            if i == 0:
                celda_label.font = fuente_encabezado
                celda_label.fill = relleno_encabezado
                celda_valor.font = fuente_encabezado
                celda_valor.fill = relleno_encabezado
            else:
                celda_label.font = fuente_datos
                celda_valor.font = fuente_datos
                celda_valor.alignment = alineacion_centro

        # Distribución por categoría
        fila_cat = 3 + len(resumen_datos) + 1
        ws_resumen.cell(row=fila_cat, column=1, value='Distribución por Categoría').font = Font(
            name='Calibri', size=11, bold=True, color='1A3A5C'
        )

        fila_cat += 1
        cat_encabezados = ['Categoría', 'Cantidad']
        for col, enc in enumerate(cat_encabezados, 1):
            celda = ws_resumen.cell(row=fila_cat, column=col, value=enc)
            celda.font = fuente_encabezado
            celda.fill = relleno_encabezado
            celda.border = borde_delgado

        for cat_data in resumen['por_categoria']:
            fila_cat += 1
            ws_resumen.cell(row=fila_cat, column=1, value=cat_data['categoria']).border = borde_delgado
            celda_cant = ws_resumen.cell(row=fila_cat, column=2, value=cat_data['cantidad'])
            celda_cant.border = borde_delgado
            celda_cant.alignment = alineacion_centro

        ws_resumen.column_dimensions['A'].width = 25
        ws_resumen.column_dimensions['B'].width = 15

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fecha_archivo = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f'reporte_{tipo}_{fecha_archivo}.xlsx'

    return send_file(
        buffer,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
